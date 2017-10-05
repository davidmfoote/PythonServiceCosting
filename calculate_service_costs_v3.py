import openpyxl
import networkx as nx
import pprint
import logging

#logging.basicConfig(level=logging.DEBUG)
logging.basicConfig(level=logging.WARNING)


#This consumes a file from Sharepoint Service Portfolio site
def build_services_dict(xlsxfile):
    logging.info("Building services dict from xlsx file")    
    #go read workbook
    wb = openpyxl.load_workbook(xlsxfile, read_only=True)

    #sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    sheet = wb.active
    services = {}
    parser = {"service_col": '', "type_col": '', "predecessors_col": '', "cost_col": ''}

    #figure out what columns our fields are in
    #probably 
    for header in sheet[1]:
        if header.value.lower() == 'service':
            parser["service_col"] = header.column
        elif header.value.lower() == 'service type':
            parser["type_col"] = header.column
        elif header.value.lower() == 'service dependencies':
            parser["predecessors_col"] = header.column
        elif header.value.lower() == 'cost':
            parser["cost_col"] = header.column


    #loop through the rows
    for row in range(2, sheet.max_row+1):
        #look at predecessors string, clean it up and make it an array
        predecessors_array = []
        ds = (sheet.cell(row=row, column=parser['predecessors_col']).value).split(';')
        for _ in ds:
            _ = _.strip('#')
            if not _.isnumeric():
                predecessors_array.append(_) 
        #build dict with relevant values
        name = sheet.cell(row=row, column=parser['service_col']).value
        serdict = {        
                "type": sheet.cell(row=row, column=parser['type_col']).value,
                "predecessors": predecessors_array,
                "base_cost": sheet.cell(row=row, column=parser['cost_col']).value,
                "location": '',
                "total_cost_from_predecessors": 0,
                "costs_from_predecessors": {},
                "cost_to_successors": None,
                "cost_to_successors_dict": {},
                "number_of_successors": 0,
                "total_cost": 0,
                "level": 0,
                "circular_relationship": [],
                "account_for_all_predecessors_costs": '',
                "number_of_successors_no_cost_pass": 0,
                }
        services.update({ name: serdict})

    return services

def build_services_digraph(services):
    logging.info("Building digraph from services dict")
    G = nx.DiGraph()

    #add nodes
    for service in services:
        G.add_node(service)

    #add edges
    for service in services:
        for ind_predecessors in services[service]['predecessors']:
            if ind_predecessors != '': 
               G.add_edge(ind_predecessors, service, {'cost': services[ind_predecessors]['cost_to_successors']})

    return G



def set_location_and_root(services, G):
    logging.info("Setting location by service")
    root = []
    for n in G.nodes():
        if len(G.successors(n)) == 0 and len(G.predecessors(n)) != 0:
            services[n]['location'] = 'top'
        elif len(G.predecessors(n)) == 0 and len(G.successors(n)) != 0:
            services[n]['location'] = 'root'
            root.append(n)
        elif len(G.successors(n)) == 0 and len(G.predecessors(n)) == 0:
            services[n]['location'] = 'alone'
            services[n]['total_cost'] = services[n]['base_cost']
        else:
            services[n]['location'] = 'middle'
    if len(root) == 1:
        return services, root[0]
    else:
        raise Exception("Too many roots")

def find_cicular_predecessors(services, G):
    #lets find where circular relationships exist in the 
    for (parent, child, data) in G.edges_iter(data=True):
        if G.has_edge(child, parent):
            logging.info("Circular relationship found between parent %s and child %s", parent, child)
            services[child]['circular_relationship'].append(parent) 
    return services

def calculate_level(services, G):
    logging.info("calculating levels")
    logging.debug("pydot needs graphviz installed and accessible via user path")
    #https://stackoverflow.com/questions/13938770/how-to-get-the-coordinates-from-layout-from-graphviz
    #build a dot graph using graphviz and then graph the y coordinate to figure out services position in hierarchy relative
    #to each other.  A hell of a lot cleaner than trying to do it myself (although it does require graphviz to be installed
    #locally).
    #root is at top, so lower numbers means further along the tree/more dependencies
    pos = nx.nx_pydot.graphviz_layout(G, prog = 'dot')
    for s in pos:
        (x, y) = pos[s]
        services[s]['level'] = y     
    return services

def calculate_node_costs(stack, services, G):
    logging.info("Calculating Node costs")
    while stack: 
        current_node = stack.pop(0)
        logging.debug("---------------Now working on: %s ---------------", current_node)
        logging.debug("current state of stack: %s", len(stack))
        current_service = services[current_node]

        #clear values in case a service got put back on the queue
        current_service['total_cost_from_predecessors'] = 0
        current_service['costs_from_predecessors'] = {}

        
        #has no children, so just set base
        if current_service['location'] == 'root' or current_service['location'] == 'alone':
            logging.debug("At root or alone - simple calc")
            current_service["total_cost"] = current_service['base_cost']
        else:
            logging.debug("looking at %s predecessors", len(G.predecessors(current_node)))
            #loop through predecessors to tally up costs
            for parent in G.predecessors(current_node):
                if current_service['level'] > services[parent]['level'] and parent in current_service['circular_relationship']:
                    logging.debug("set cost of parent %s  to 0 because it's higher level and in a circular relationship", parent)
                    current_service['costs_from_predecessors'][parent]  = 0
                    services[parent]['cost_to_successors_dict'][current_node] = 0
                    services[parent]['number_of_successors_no_cost_pass'] = services[parent]['number_of_successors_no_cost_pass'] +1
                elif services[parent]['cost_to_successors']:
                    current_service['total_cost_from_predecessors'] = current_service['total_cost_from_predecessors'] + services[parent]['cost_to_successors']
                    current_service['costs_from_predecessors'][parent]  = services[parent]['cost_to_successors']
                    services[parent]['cost_to_successors_dict'][current_node] = services[parent]['cost_to_successors']
                    logging.debug("cost from %s was %s", parent, services[parent]['cost_to_successors'])
                #elif parent in current_service['circular_relationship']:
                #      logging.debug("probably shouldn't be here")  
                else:
                    #found a parent that was not fully calculated
                    logging.debug(" %s not fully calculated, skipping calculation, adding %s back on to the stack to be recalculated", parent, current_node)
                    stack.append(current_node)
                    
            logging.debug("finished with parents")
            current_service["total_cost"] = int(current_service['base_cost']) + int(current_service['total_cost_from_predecessors'])

        ##Here is where we would set the different costs
        if G.successors(current_node):
            if current_service['type'] == 'IT Supporting':
                current_service['cost_to_successors'] = int(current_service['total_cost'] / len(G.successors(current_node)))
            else:
                logging.debug("%s is Business Facing, set multipler for pass on costs to be .1", current_node)
                current_service['cost_to_successors'] = int((int(current_service['total_cost'] / len(G.successors(current_node)))) * .1)
        current_service['number_of_successors'] = len(G.successors(current_node))
        current_service["account_for_all_predecessors_costs"] = (len(current_service['predecessors']) == len(current_service["costs_from_predecessors"]))
        logging.debug("Base cost was %s and took on from parents %s", current_service['base_cost'],current_service['total_cost_from_predecessors'] )
        logging.debug("There are %s successors: %s", len(G.successors(current_node)), G.successors(current_node))
        logging.debug("Total cost is %s divided by %s and pass on cost is %s", current_service['total_cost'],len(G.successors(current_node)), current_service['cost_to_successors'])
        
        
    return services

def output_graphviz(services, graphviz_file):
    logging.info("Creating graphviz file at %s", graphviz_file)
    newG = build_services_digraph(services)

    with open(graphviz_file, "w") as out_file:
        print('//https://stamm-wilbrandt.de/GraphvizFiddle \ndigraph services {\n    ratio = fill;\n    node [style=filled];\n    size = "15, 25";', file=out_file)
        for node in newG.nodes():
            if services[node]['type'] == 'IT Supporting':
                print('    "{}" [fillcolor=lightgrey label="{}\\n${}"]'.format(node, node, services[node]['total_cost']), file=out_file)
            else:
                print('    "{}" [fillcolor=green3 label="{}\\n${}"]'.format(node, node, services[node]['total_cost']), file=out_file)

        for service in services:
            for pred in services[service]['costs_from_predecessors']:
                print('    "{}" -> "{}" [label="${}"]'.format(pred, service, services[service]['costs_from_predecessors'][pred]), file=out_file)
        print('}', file=out_file)
    return


def output_excel(services, dest_filename):
    #The point of this is to create an excel sheet for each service
    #that uses forumlas to link successor and predecessor services
    #It will intially use values that the calculate_services_costs.py calculates
    #but will allow for changes made to the excel to cascade down the service structure
    #It also allows for greater control of costs passed to successors.  Even though it starts
    #by using the straight percentage forumla, by having each item listed, the individual successor
    #costs can be changed.

    #excel parser
    #Store where we want to put various pieces of information
    logging.info("Creating excel file at %s", dest_filename)
    loc = {
        #because we iterate over rows, it's easiest to use openpyxl's row,column way of specifying cells
        #so listing column values here rather than column letters is appropriate
        #however, excel formulas like column letters, so we create a spot for them and then use openpyxl
        #to calculate and store those as well
        'precessor_service_col' : 5,
        'precessor_service_col_letter' : '',
        'precessor_value_col' : 6,
        'precessor_value_col_letter' : '',
        'successor_sum_col' : 1,
        'successor_sum_col_letter' : '',
        'successor_value_col' : 2,
        'successor_value_col_letter' : '',
        'successor_per_col' : 3,
        'successor_per_col_letter' : '',
        'successor_services_row_start' : 11,
        'precessor_services_row_start' : 3,
        
        'service_title': 'A1',
        'base_cost_label': 'A2',
        'base_cost' : 'C2',
        'number_of_successors' : 'C7',
        'cost_from_predecessors_label': 'A3',
        'total_cost_from_predecessors' : 'C3',
        'total_cost' : 'C5',
        'cost_to_successors' : 'C8',
        'type_name' : 'B6',
        'type_factor' : 'C6',
                      
        'total_cost_label': 'A5',
        'type_label': 'A6',
        'num_successors_label': 'A7',
        'cost_successors_label': 'A8',
        'list_of_successors_label': 'A10',
        'percentages_instructions_label': 'B10',
        
        
        }

    loc['precessor_service_col_letter'] = openpyxl.utils.get_column_letter(loc['precessor_service_col'])
    loc['precessor_value_col_letter'] = openpyxl.utils.get_column_letter(loc['precessor_value_col'])
    loc['successor_sum_col_letter'] = openpyxl.utils.get_column_letter(loc['successor_sum_col'])
    loc['successor_value_col_letter'] = openpyxl.utils.get_column_letter(loc['successor_value_col'])
    loc['successor_per_col_letter'] = openpyxl.utils.get_column_letter(loc['successor_per_col'])

    wb = openpyxl.Workbook()

    #Since excel complains about sheet names being too long
    #find names over 30 and turn them into ugly abbreviations
    sheet_names = {}
    for service in services:
        if len(service) > 30:
            sn = "".join(e[0:3] for e in service.split())
        else:
            sn = service
        sheet_names[service] = sn

    #Build the excel sheet for a given service
    for service in services:
        sheet_name = sheet_names[service]
        wb.create_sheet(title=sheet_name)
        ws = wb.get_sheet_by_name(sheet_name)
        #This section is pulling out items that don't need to be calculated
        #like base cost, number of successors, etc
        ws[loc['service_title']] = service
        ws[loc['service_title']].style = 'Title'

        ws[loc['base_cost_label']] = 'Base Cost'
        ws[loc['base_cost']] = services[service]['base_cost']

        #This the type factor (to make Business Facing cost less for successors)
        #it is baked into the calculate services cost python, but we need to recreated it here
        #in case any values change
        ws[loc['type_label']] = 'Type'
        ws[loc['type_name']] = services[service]['type']
        if services[service]['type'] == 'IT Supporting':
            ws[loc['type_factor']] = 1
        else:
            ws[loc['type_factor']] = .1
        
        ws[loc['num_successors_label']] = 'Number of successors'
        ws[loc['number_of_successors']] = services[service]['number_of_successors'] - services[service]['number_of_successors_no_cost_pass']
        
        ws[loc['cost_successors_label']] = 'Cost to successors'
                #The forumla for total cost to successors is calculated by python and sorted in the services dict
        #but to make sure it changes if we edit the excel, use a formula like =IFERROR((C5/C7)*C6,0)
        ws[loc['cost_to_successors']] = '=IFERROR((' + loc['total_cost'] + '/' + loc['number_of_successors'] + ')*' + loc['type_factor'] + ',0)'
        ws[loc['cost_to_successors']].style = 'Calculation'

        #List all the costs from the predecessors
        if len(services[service]['costs_from_predecessors']) > 0:
            ws.cell(row=2, column=loc['precessor_service_col'], value='Cost from Predecessors')
            r = loc['precessor_services_row_start']
            for pred in services[service]['costs_from_predecessors']:
                v = services[service]['costs_from_predecessors'][pred]
                #put the name of the predecessor
                ws.cell(row=r, column=loc['precessor_service_col'], value=pred)
                
                #prepare a formula to grab value from excel
                #excel isn't smart enough to know which is above or below, so it we aren't counting cost
                #in python, then just put a zero
                if v != 0:
                    #predecessor_cost_formula = "='" + sheet_names[pred] + "'!"+ loc['cost_to_successors']
                    predecessor_cost_formula = "=VLOOKUP(" + loc['successor_sum_col_letter'] + "1, '" + sheet_names[pred] + "'!" + loc['successor_sum_col_letter'] + str(loc['successor_services_row_start']) + ":" + loc['successor_value_col_letter'] + "100,2, FALSE)"
                else:
                    predecessor_cost_formula = 0
                ws.cell(row=r, column=loc['precessor_value_col'], value=predecessor_cost_formula).style = 'Linked Cell'
                ws.cell(row=r, column=loc['precessor_value_col']).number_format =  '$#,##0'
                r += 1

            pred_formula = "=sum("+ loc['precessor_value_col_letter'] + str(loc['precessor_services_row_start']) + ":" + loc['precessor_value_col_letter'] + str(r-1) + ")"
            pred_sum = ws.cell(row=r, column=loc['precessor_value_col'], value=pred_formula)
            pred_sum.style = 'Calculation'
            pred_sum.number_format = '$#,##0'
            #calculate sum of precedessors
            ws.cell(row=r, column=loc['precessor_service_col'], value="Predecessors Sum").style = 'Total'

        #set successor values
        if services[service]['number_of_successors'] > 0:
            ws[loc['list_of_successors_label']] = 'Cost to Successors'
            ws[loc['percentages_instructions_label']] = 'Change the percentages below to change what is passed on'
            sr=loc['successor_services_row_start']
            for succ in services[service]['cost_to_successors_dict']:
                v = services[service]['cost_to_successors_dict'][succ]
                num_suc = services[service]['number_of_successors'] - services[service]['number_of_successors_no_cost_pass']
                ws.cell(row=sr, column=loc['successor_sum_col'], value=succ)
                #if successor cost is more than zero
                if v != 0:
                    #For this formula to work, we need to enable Iterative Calculations
                    #Even then though, Excel really cannot figure this formula out if we point directly to Total Cost - I have no idea why -
                    #but if we recalculate total cost by taking the already calculated Cost to Successors times the number of them, it works, somehow
                    per_formula = "=" + "(" + loc['cost_to_successors'] + "*" + loc['number_of_successors'] + ")" + "*"  + loc['successor_per_col_letter'] + str(sr)
                    ws.cell(row=sr, column=loc['successor_value_col'], value=per_formula)
                    ws.cell(row=sr, column=loc['successor_value_col']).number_format = '$#,##0'
                    ws.cell(row=sr, column=loc['successor_per_col'], value= 1 / num_suc)
                    ws.cell(row=sr, column=loc['successor_per_col']).style = 'Percent'
                #add a comment if a zero is passed to a successor
                else:
                    ws.cell(row=sr, column=loc['successor_per_col'], value=0)
                    ws.cell(row=sr, column=loc['successor_per_col']).style = 'Percent'
                    comment = openpyxl.comments.Comment('Do not pass costs to this successor because is below', 'Script')
                    ws.cell(row=sr, column=loc['successor_value_col']).comment = comment
                    number_comment = openpyxl.comments.Comment('Excluding successors who do not share cost here', 'Script')
                    ws[loc['number_of_successors']].comment = number_comment
                    
                sr += 1

            succ_formula = "=sum(" + loc['successor_value_col_letter'] + str(loc['successor_services_row_start']) + ":" + loc['successor_value_col_letter'] + str(sr-1) + ")"
            succ_sum = ws.cell(row=sr, column=loc['successor_value_col'], value=succ_formula)
            succ_sum.style = 'Calculation'
            succ_sum.number_format = '$#,##0'
            per_sum_forumla = "=sum(" + loc['successor_per_col_letter'] + str(loc['successor_services_row_start']) + ":" + loc['successor_per_col_letter'] + str(sr-1) + ")"
            per_sum = ws.cell(row=sr, column=loc['successor_per_col'], value=per_sum_forumla)
            per_sum.style = 'Percent'
            ws.cell(row=sr, column=loc['successor_sum_col'], value="Successor Sum").style = 'Total'
            
        

        
        ws[loc['cost_from_predecessors_label']] = 'Costs from Predecessors'
        ws[loc['total_cost_from_predecessors']] = '='+ pred_sum.coordinate
        ws[loc['total_cost_from_predecessors']].style = 'Calculation'

        ws[loc['total_cost_label']] = 'Total Cost'
        ws[loc['total_cost']] = '=' + loc['base_cost'] + '+' + loc['total_cost_from_predecessors']
        ws[loc['total_cost']].style = 'Calculation'

        ws[loc['base_cost']].number_format = '$#,##0'
        
        ws[loc['total_cost_from_predecessors']].number_format = '$#,##0'
        ws[loc['total_cost']].number_format = '$#,##0'
        ws[loc['cost_to_successors']].number_format = '$#,##0'


    #Change the default sheet to reference the total cost values
    #of each service
    toc = wb.get_sheet_by_name('Sheet')
    toc.title = 'Totals'
    #r = 1
    for r, service in enumerate(services, 1):
        toc.cell(row=r, column=1, value=service)
        service_total_formula = "='" + sheet_names[service] + "'!"+ loc['total_cost']
        toc.cell(row=r, column=2, value=service_total_formula).number_format = '$#,##0'
        #r += 1


    toc.cell(row=1, column=4, value="If you see an error upon opening, please enable Iterative Formulas by going to File->Options->Formulas").style = 'Explanatory Text'
    
    #save excel
    
    wb.save(filename = dest_filename )

    return


if __name__=='__main__':
    #This is from the Service Portfolio list on Sharepoint
    #Expects Service, Service Type, Service Dependencies, and Cost
    #Currently Service Dependencies is a semicolon separated list of values

    xlsxfile = 'services.xlsx' #full file

    services = build_services_dict(xlsxfile)
    G = build_services_digraph(services)
    services, root = set_location_and_root(services, G)
    services = find_cicular_predecessors(services, G)
 
    services = calculate_level(services, G)
    
    #build list of levels, so we can put services in order
    num_levels = []
    for s in services:
        num_levels.append(services[s]['level'])
    num_levels = sorted(list(set(num_levels)), reverse=True)
    
    #build stack based on level of service
    stack = []
    for current_level in num_levels:
        new_level = []
        for s in services:
            if services[s]['level'] == current_level:
                stack.append(s)

    services = calculate_node_costs(stack, services, G)   
    output_graphviz(services, 'graphviz.txt')
    output_excel(services, 'services_costs.xlsx')
    
    #temp error checking
    for s in services:
        if services[s]['location'] != 'root' and services[s]['location'] != 'alone':
            if services[s]['account_for_all_predecessors_costs'] != True:
                logging.warning("%s does not account for all predecessor costs", s)
    
